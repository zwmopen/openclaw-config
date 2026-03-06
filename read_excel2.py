import openpyxl
import os
import sys

# 设置UTF-8输出
sys.stdout.reconfigure(encoding='utf-8')

base = r'D:\Program Files\Obsidian\zwm\.zwm\个人知识库'
found_files = []

for root, dirs, files in os.walk(base):
    for f in files:
        if f.endswith('.xlsx'):
            full_path = os.path.join(root, f)
            try:
                wb = openpyxl.load_workbook(full_path, read_only=True)
                sheet = wb.active
                # 检查是否有"碎片"或"高版本"相关内容
                has_fragment = False
                for row in range(1, min(10, sheet.max_row+1)):
                    for col in range(1, min(10, sheet.max_column+1)):
                        cell_val = sheet.cell(row=row, column=col).value
                        if cell_val and ('碎片' in str(cell_val) or '高版本' in str(cell_val) or '强者' in str(cell_val)):
                            has_fragment = True
                            break
                    if has_fragment:
                        break
                wb.close()
                
                if has_fragment:
                    found_files.append(full_path)
            except Exception as e:
                pass

print(f'找到 {len(found_files)} 个文件包含碎片/强者内容:')
for path in found_files:
    print(f'\n{"="*80}')
    print(f'文件: {path}')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    print(f'Sheet: {sheet.title}, 行数: {sheet.max_row}, 列数: {sheet.max_column}')
    print('-'*80)
    
    # 打印所有内容
    for row in range(1, min(101, sheet.max_row+1)):
        cells = []
        for col in range(1, min(8, sheet.max_column+1)):
            val = sheet.cell(row=row, column=col).value
            if val:
                cells.append(str(val))
        if cells:
            print(f'第{row}行: {" | ".join(cells)}')
    wb.close()
